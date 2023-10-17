import itertools

import openpyxl


class Record:
    def __init__(self,
        blogger_nickname, blogger_homepage, link, title, publication_date,
        body, number_of_likes, number_of_favorites, number_of_comments,
        comment_1, comment_2, comment_3, comment_4, comment_5, comment_6):
        self.blogger_nickname = blogger_nickname
        self.blogger_homepage = blogger_homepage
        self.link = link
        self.title = title
        self.publication_date = self.clean_date(publication_date)
        self.body = body
        self.number_of_likes = self.clean_number(number_of_likes)
        self.number_of_favorites = self.clean_number(number_of_favorites)
        self.number_of_comments = self.clean_number(number_of_comments)
        self.number_of_interactions = self.number_of_likes + self.number_of_favorites + self.number_of_comments
        self.comment_1 = comment_1
        self.comment_2 = comment_2
        self.comment_3 = comment_3
        self.comment_4 = comment_4
        self.comment_5 = comment_5
        self.comment_6 = comment_6
    
    @staticmethod
    def clean_number(number):
        number = str(number)
        number = number.strip()
        number = number.rstrip('+')
        
        factor = 1
        if number.endswith('w'):
            factor = 10000
            number = number[:-1]
        elif number.endswith('k'):
            factor = 10000
            number = number[:-1]

        number = float(number)
        number = number * factor
        number = int(number)
        return number

    @staticmethod
    def clean_date(date):
        date = str(date)
        date = date.strip()
        date = date[:10]
        return date

    def __eq__(self, other):
        return self.link == other.link

    def __hash__(self) -> int:
        return hash(self.link)


def main():
    src_wb = openpyxl.load_workbook('src1.xlsx', data_only=True, rich_text=True)
    src_sheet = src_wb.active

    # Get the column names and indices
    columns = {cell.value: cell.column - 1 for cell in next(src_sheet.iter_rows())}

    records = set()
    for row in itertools.islice(src_sheet.iter_rows(values_only=True), 1, None):
        record = Record(
            blogger_nickname=row[columns['博主昵称']],
            blogger_homepage=row[columns['博主主页']],
            link=row[columns['链接']],
            title=row[columns['标题']],
            publication_date=row[columns['发布日期']],
            body=row[columns['正文']],
            number_of_likes=row[columns['点赞数']],
            number_of_favorites=row[columns['收藏数']],
            number_of_comments=row[columns['评论数']],
            comment_1=row[columns['评论1']],
            comment_2=row[columns['评论2']],
            comment_3=row[columns['评论3']],
            comment_4=row[columns['评论4']],
            comment_5=row[columns['评论5']],
            comment_6=row[columns['评论6']],
        )
        records.add(record)
    
    clean_wb = openpyxl.Workbook()
    clean_sheet = clean_wb.active

    clean_sheet.append([
        '博主昵称', '博主主页', '链接', '标题', '发布日期',
        '正文', '点赞数', '收藏数', '评论数', '互动数',
        '评论1', '评论2', '评论3', '评论4', '评论5', '评论6',
    ])
    for record in records:
        row = [
            record.blogger_nickname,
            record.blogger_homepage,
            record.link,
            record.title,
            record.publication_date,
            record.body,
            record.number_of_likes,
            record.number_of_favorites,
            record.number_of_comments,
            record.number_of_interactions,
            record.comment_1,
            record.comment_2,
            record.comment_3,
            record.comment_4,
            record.comment_5,
            record.comment_6,
        ]
        clean_sheet.append(row)

    clean_wb.save('clean.xlsx')

main()
